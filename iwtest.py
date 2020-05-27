import csv
import os
import json
import urllib.parse
import urllib.request
import ssl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side

HTTP_HEADERS = {'X-API-Version': '1.2',
                'X-API-CompanyId': 'GUAP',
                'X-API-ImporterName': 'GUAP',
                'X-API-Auth-Token': '1bs23q0mf47941ctode8'}
ip = '10.228.6.236:17443'
url = 'https://' + ip + '/xapi/event/protectedDocument?with[conditions]start=0&limit=10'
req = urllib.request.Request(url, headers=HTTP_HEADERS)
data = []
with urllib.request.urlopen(req, context=ssl._create_unverified_context()) as response:
    the_page = response.read()
    data = json.loads(the_page)

with open('iw-test-data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

book = Workbook()
sheet = book.active

sheet['A1'] = 'Политика 1'
sheet['A2'] = 'Политика 2'
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

sheet.cell(4, 4).value = 'blalbla'
sheet.cell(4, 4).border = thin_border
sheet.cell(4, 5).value = 'blal'

book.save("sample.xlsx")