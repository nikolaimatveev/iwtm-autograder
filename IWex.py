# -*- coding: utf-8 -*-

import urllib.parse
import urllib.request
import ssl
from pprint import pprint
import json
from openpyxl import load_workbook
import sys
from openpyxl.styles import Font, Color, colors
from datetime import datetime
import calendar
from openpyxl.styles import PatternFill
#import LDAP


class WrongInputException(Exception):
    pass


NUMBER_OF_EVENTS = 1000
LIST_OF_EVENTS = {}
# Здесь описываются данные, необходимые для соединения с IWTM
X_API_Version = 1
X_API_CompanyId = "GUAP"
X_API_ImporterName = "GUAP"
ip_addr = 0
HTTP_HEADERS = {'X-API-Version': X_API_Version,
                'X-API-CompanyId': X_API_CompanyId,
                'X-API-ImporterName': X_API_ImporterName,
                'X-API-Auth-Token': ""}
users_js = {}


# Помещение событий, полученных из запроса в словарь с ключом OBJECT_ID
def create_list_of_events(data):
    for i in range(len(data['data'])):
        event = data['data'][i]
        # if event['OBJECT_ID'] not in LIST_OF_EVENTS and len(event['policies']) != 0:
        if event['OBJECT_ID'] not in LIST_OF_EVENTS:
            LIST_OF_EVENTS[event['OBJECT_ID']] = event


# Получить имя файла из события
def get_name_from_event(event_id):
    url = 'https://' + ip_addr + '/xapi/event/' + str(event_id) + '/raw?downoload=0'
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, context=ssl._create_unverified_context()) as response:
        the_page = response.read()
        try:
            if len(the_page) > 390:
                meta = the_page[0:390].decode()
                start = meta.find("filename=")
                if (start == -1):
                    filename = ""
                else:
                    end = meta.find('\"', start + 10)
                    filename = meta[start + 10:end]
            else:
                filename = ""
        except Exception as e:
            filename = ""
        return filename


# Получение событий в формате JSON от IWTM
def get_events_from_url(url):
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, context=ssl._create_unverified_context()) as response:
        the_page = response.read()
        data = json.loads(the_page)
    
    with open('data0505.json', 'w', encoding='utf-8') as outfile:
        json.dump(data, outfile, ensure_ascii=False)
    return data


# Сохранение файла, нарушающего политику
def save_event_file(url):
    filename = "event_file.txt"
    file = open(filename, 'wb')
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, context=ssl._create_unverified_context()) as response:
        the_page = response.read()
        file.write(the_page)
        file.close()


# Окрашивание ячейки
def fill_cell(cell, template_value, fact_value, warning):
    cell.value = fact_value
    if warning == True:
        cell.value = fact_value + " (" + template_value + ")"
        cell.fill = PatternFill(patternType='solid',
                                fill_type='solid',
                                fgColor=Color('FFFF00'))
        return
    if template_value == fact_value:
        cell.fill = PatternFill(patternType='solid',
                                fill_type='solid',
                                fgColor=Color('00FF00'))
    else:
        cell.value = fact_value + " (" + template_value + ")"
        cell.fill = PatternFill(patternType='solid',
                                fill_type='solid',
                                fgColor=Color('FF0000'))


# Заполнить шаблон Факт
def save_fact_to_table(template_table, fact_table, wb):
    sheet = wb['Факт']
    column_number = 4
    for i in range(len(fact_table)):
        cell = sheet['A' + str(column_number)]
        if template_table[i]['tags'] == 'none':
            template_table[i]['tags'] = ""
        if template_table[i]['policy'] == 'none':
            template_table[i]['policy'] = ""
        # print(template_table[i]['policy'] + " " + fact_table[i]['policy'])
        cell.value = fact_table[i]['task']
        cell = sheet['B' + str(column_number)]
        cell.value = fact_table[i]['event_number']
        cell = sheet['C' + str(column_number)]
        fill_cell(cell, template_value=template_table[i]['file_name'], fact_value=fact_table[i]['file_name'],
                  warning=False)
        cell = sheet['D' + str(column_number)]
        fill_cell(cell, template_value=template_table[i]['sender'], fact_value=fact_table[i]['sender'], warning=False)
        cell = sheet['E' + str(column_number)]
        fill_cell(cell, template_value=template_table[i]['recipient'], fact_value=fact_table[i]['recipient'],
                  warning=False)
        cell = sheet['F' + str(column_number)]
        policies = ""
        warning = False
        if len(fact_table[i]['policies']) == 1:
            policies = fact_table[i]['policies'][0]
        else:
            for j in range(0, len(fact_table[i]['policies'])):
                if fact_table[i]['policies'][j] == template_table[i]['policy'] and len(fact_table[i]['policies']) > 1:
                    warning = True
                policies += fact_table[i]['policies'][j] + " "
        fill_cell(cell, template_value=template_table[i]['policy'], fact_value=policies, warning=warning)
        cell = sheet['G' + str(column_number)]
        fill_cell(cell, template_value=template_table[i]['verdict'], fact_value=fact_table[i]['verdict'], warning=False)
        cell = sheet['H' + str(column_number)]
        fill_cell(cell, template_value=template_table[i]['violation_level'],
                  fact_value=fact_table[i]['violation_level'], warning=False)
        cell = sheet['I' + str(column_number)]
        tags = ""
        warning = False
        if len(fact_table[i]['tags']) == 1:
            tags = fact_table[i]['tags'][0]
        else:
            for j in range(0, len(fact_table[i]['tags'])):
                if fact_table[i]['tags'][j] == template_table[i]['tags'] and len(fact_table[i]['tags']) > 1:
                    warning = True
                tags += fact_table[i]['tags'][j] + " "
        fill_cell(cell, template_value=template_table[i]['tags'], fact_value=tags, warning=warning)
        cell = sheet['J' + str(column_number)]
        cell.value = fact_table[i]['date']
        cell = sheet['K' + str(column_number)]
        cell.value = fact_table[i]['ID']
        cell = sheet['L' + str(column_number)]
        cell.value = fact_table[i]['protected_documents']
        column_number += 1


def table_comparison(input_table_name, output_table_name):
    wb = load_workbook(filename=input_table_name)
    users_js.__delitem__('External')
    column_number = 4
    sheet = wb['Шаблон']
    template_table = []
    while (True):
        column = {}
        column['task'] = str(sheet['A' + str(column_number)].value).lower()
        column['event_number'] = sheet['B' + str(column_number)].value
        if column['event_number'] == None:
            break
        column['file_name'] = sheet['C' + str(column_number)].value
        column['sender'] = sheet['D' + str(column_number)].value
        column['recipient'] = sheet['E' + str(column_number)].value
        column['policy'] = str(sheet['F' + str(column_number)].value).lower()
        column['verdict'] = str(sheet['G' + str(column_number)].value.lower()).lower()
        column['violation_level'] = str(sheet['H' + str(column_number)].value.lower()).lower()
        column['tags'] = str(sheet['I' + str(column_number)].value).lower()
        template_table.append(column)
        column_number += 1
    fact_table = []
    finded = 0
    for item in template_table:
        for event_id in LIST_OF_EVENTS:
            column = {}
            event = LIST_OF_EVENTS[event_id]
            """
            preview_data = json.loads(str(LIST_OF_EVENTS[event_id]['PREVIEW_DATA']))
            if 'name' not in preview_data:
                column['file_name'] = ""
            else:
                column['file_name'] = preview_data['name']
                """
            column['file_name'] = get_name_from_event(event_id)
            column['sender'] = []
            if len(LIST_OF_EVENTS[event_id]['senders']) == 0:
                column['sender'].append('External')
            else:
                for i in range(0, len((LIST_OF_EVENTS[event_id]['senders'])[0][0]['keys'])):
                    if LIST_OF_EVENTS[event_id]['senders'][0][0]['keys'][i]['KEY_TYPE'].lower() == "email":
                        sender_email = (LIST_OF_EVENTS[event_id]['senders'])[0][0]['keys'][i]['KEY'].lower()
                # print(sender_full_name)
                for group in users_js:
                    for people in users_js[group]:
                        for human in people:
                            # print(people[human])
                            email = people[human]['email'].decode().split(" ")[0].lower()
                            if len(email) > 1:
                                # last_name = full_name[1].replace('\'', "")
                                # first_name = full_name[0].replace('\'', "")
                                # if LIST_OF_EVENTS[event_id]['OBJECT_ID'] == 285:
                                # if sender_full_name[0] == first_name or sender_full_name[0] == last_name:
                                # print(sender_full_name[0] + "==" + first_name + " " + sender_full_name[2] + "==" + last_name)
                                # print(sender_full_name)
                                # print(name)
                                # if column['file_name'] == "g1.rtf" and sender_full_name[0]==last_name:
                                # print(sender_full_name[0] + "==" + last_name + " " + sender_full_name[2] + "==" + first_name)
                                # print(sender_full_name[0] + "==" + last_name + " " + sender_full_name[2] + "==" + first_name)
                                if sender_email == email:
                                    column['sender'].append(group)
            column['recipient'] = []
            if len(LIST_OF_EVENTS[event_id]['recipients']) == 0:
                column['recipient'].append('External')
            else:
                for i in range(0, len((LIST_OF_EVENTS[event_id]['recipients'])[0][0]['keys'])):
                    if LIST_OF_EVENTS[event_id]['recipients'][0][0]['keys'][i]['KEY_TYPE'].lower()=="email":
                        recipient_email = (LIST_OF_EVENTS[event_id]['recipients'])[0][0]['keys'][i]['KEY'].lower()
                for group in users_js:
                    for people in users_js[group]:
                        for human in people:
                            email = people[human]['email'].decode().split(" ")[0].lower()
                            if len(email) > 1:
                                # last_name = full_name[1].replace('\'', "")
                                # first_name = full_name[0].replace('\'', "")
                                # if LIST_OF_EVENTS[event_id]['OBJECT_ID'] == 61:
                                # print(recipient_full_name[0] + "==" + first_name + " " + recipient_full_name[1] + "==" + last_name)
                                if recipient_email == email:
                                    column['recipient'].append(group)
            # if item['file_name'] == "win1.zip" and column['file_name'] == "win1.zip":
            #    print(item['file_name'] + " [] " + column["file_name"] + " [] " + item['sender'] + " [] " + column['sender'] + " [] " + item['recipient'] + " [] " + column['recipient'])
            # if item['file_name'] == "product3.rtf" or item['file_name'] == "win1.zip" or item['file_name'] == "doc1.rtf":
            #    print(item['file_name'] + " [] " + column["file_name"] + " [] " + item['sender'] + " [] " + column['sender'] + " [] " + item['recipient'] + " [] " + column['recipient'])
            # if item['file_name'] == "g1.rtf" and column['file_name'] == "g1.rtf":
            # print(item)
            # print(column)
            # print(str(LIST_OF_EVENTS[event_id]['OBJECT_ID']) + ":" +  item['file_name'] + " [] " + column["file_name"] + " [] " + item['sender'] + " [] " + column['sender'] + " [] " + item['recipient'] + " [] " + column['recipient'])
            right_sender = 0
            sender_group = "none"
            for i in range(0, len(column['sender'])):
                if column['sender'][i] == item['sender']:
                    right_sender = 1
                    sender_group=column['sender'][i]
            column['sender']=sender_group
            right_recipient = 0
            recipient_group = "none"
            for i in range(0, len(column['recipient'])):
                if column['recipient'][i] == item['recipient']:
                    right_recipient = 1
                    recipient_group = column['recipient'][i]
            column['recipient']=recipient_group
            if item['file_name'] == column["file_name"] and right_sender == 1 and right_recipient == 1:
                finded = 1
                column['event_number'] = item['event_number']
                column['task'] = item['task'].lower()
                if len(LIST_OF_EVENTS[event_id]['policies']) != 0:
                    list_of_policies = (LIST_OF_EVENTS[event_id]['policies'])
                    column['policies'] = []
                    for i in range(0, len(list_of_policies)):
                        name = LIST_OF_EVENTS[event_id]['policies'][i]['DISPLAY_NAME']
                        column['policies'].append(name.lower())

                else:
                    column['policies'] = []
                column['verdict'] = LIST_OF_EVENTS[event_id]['VERDICT'].lower()
                column['violation_level'] = LIST_OF_EVENTS[event_id]['VIOLATION_LEVEL'].lower()
                if len(LIST_OF_EVENTS[event_id]['tags']) != 0:
                    list_of_policies = (LIST_OF_EVENTS[event_id]['tags'])
                    column['tags'] = []
                    for i in range(0, len(list_of_policies)):
                        name = LIST_OF_EVENTS[event_id]['tags'][i]['DISPLAY_NAME']
                        column['tags'].append(name.lower())

                else:
                    column['tags'] = []

                # column['tags'] = "WorldSkills"
                column['ID'] = LIST_OF_EVENTS[event_id]['OBJECT_ID']
                column["date"] = LIST_OF_EVENTS[event_id]['CAPTURE_DATE']
                if len(LIST_OF_EVENTS[event_id]['protected_documents']) != 0:
                    list_of_protected_documents = (LIST_OF_EVENTS[event_id]['protected_documents'])
                    column['protected_documents'] = ""
                    for i in range(0, len(list_of_protected_documents)):
                        if (i == 0):
                            column['protected_documents'] += (LIST_OF_EVENTS[event_id]['protected_documents'])[i][
                                'DISPLAY_NAME']
                        else:
                            column['protected_documents'] += " " + (LIST_OF_EVENTS[event_id]['protected_documents'])[i][
                                'DISPLAY_NAME']

                else:
                    column['protected_documents'] = ""
                fact_table.append(column)
                break
        if finded == 0:
            column['event_number'] = ""
            column['policies'] = []
            column['sender'] = ""
            column['recipient'] = ""
            column['file_name'] = ""
            column['protected_documents'] = ""
            column['ID'] = ""
            column["date"] = ""
            column['tags'] = []
            column['verdict'] = ""
            column['violation_level'] = ""
            column['task'] = ""
            fact_table.append(column)
        finded = 0

    save_fact_to_table(template_table=template_table, fact_table=fact_table, wb=wb)
    wb.save(output_table_name)
    wb.close()
    print("Done!")


# Сохранение события с политикой, каталогом объекта защиты, объектом защиты в xlsx таблицу с баллами.
"""
def save_data_to_table(input_table_name, output_table_name):
    # f = open(TABLE_FILENAME, "w")
    # f.write("EVENT_ID;OBJECT_ID;DATANAME;VERDICT;USER_DECISION;VIOLATION_LEVEL;APPLICATION_FROM;POLICY_NAME;PROTECTED_DOCUMENTS_NAME;PROTECTED_CATALOGS_NAME\n")
    wb = load_workbook(filename=input_table_name)
    policy_addr = 74
    sheet = wb['CIS Marking Scheme Import']
    while (True):
        policy_cell = sheet['I' + str(policy_addr)]
        protected_catalogs_cell = sheet['J' + str(policy_addr)]
        protected_documents_cell = sheet['K' + str(policy_addr)]
        policy_type_cell = sheet['L' + str(policy_addr)]
        usages_cell = sheet['M' + str(policy_addr)]
        violation_level_cell = sheet['N' + str(policy_addr)]
        verdict_cell = sheet['O' + str(policy_addr)]
        policy = policy_cell.value
        protected_catalogs = protected_catalogs_cell.value
        protected_documents = protected_documents_cell.value
        policy_type = policy_type_cell.value
        usages = usages_cell.value
        violation_level = violation_level_cell.value
        verdict = verdict_cell.value
        if (policy == None):
            break
        recieved_usages = 0
        capture_dates = []
        first_detection = 0
        for x in LIST_OF_EVENTS:
            # name_js = json.loads(str(LIST_OF_EVENTS[x]['PREVIEW_DATA']))
            recieved_policy = LIST_OF_EVENTS[x]['policies']
            recieved_policy_name = recieved_policy[0]['DISPLAY_NAME']
            if recieved_policy_name == policy:
                capture_dates.append(
                    LIST_OF_EVENTS[x]['CAPTURE_DATE'][11:(len(LIST_OF_EVENTS[x]['CAPTURE_DATE']))] + "(" + str(
                        LIST_OF_EVENTS[x]['OBJECT_ID']) + ")")
                recieved_usages = recieved_usages + 1
                if (first_detection == 0):
                    mark = 0.2
                    recieved_protected_documents_name = ""
                    recieved_protected_catalogs_name = ""
                    if len(LIST_OF_EVENTS[x]['protected_documents']) != 0:
                        recieved_protected_documents = LIST_OF_EVENTS[x]['protected_documents']
                        recieved_protected_documents_name = recieved_protected_documents[0]['DISPLAY_NAME']
                    if len(LIST_OF_EVENTS[x]['protected_catalogs']) != 0:
                        recieved_protected_catalogs = LIST_OF_EVENTS[x]['protected_catalogs']
                        recieved_protected_catalogs_name = recieved_protected_catalogs[0]['DISPLAY_NAME']
                    recieved_violation_level = LIST_OF_EVENTS[x]['VIOLATION_LEVEL']
                    recieved_verdict = LIST_OF_EVENTS[x]['VERDICT']
                    recieved_policy_type = LIST_OF_EVENTS[x]['RULE_GROUP_TYPE']
                    if recieved_protected_catalogs_name == protected_catalogs:
                        mark = mark + 0.2
                    if recieved_protected_documents_name == protected_documents:
                        mark = mark + 0.2
                    if recieved_violation_level == violation_level:
                        mark = mark + 0.2
                    if recieved_verdict == verdict:
                        mark = mark + 0.2
                    recieved_policy_cell = sheet['Q' + str(policy_addr)]
                    recieved_protected_catalogs_cell = sheet['R' + str(policy_addr)]
                    recieved_protected_documents_cell = sheet['S' + str(policy_addr)]
                    recieved_policy_type_cell = sheet['T' + str(policy_addr)]
                    recieved_violation_level_cell = sheet['V' + str(policy_addr)]
                    recieved_verdict_cell = sheet['W' + str(policy_addr)]
                    mark_cell = sheet['X' + str(policy_addr)]
                    recieved_protected_catalogs_cell._value = str(recieved_protected_catalogs_name)
                    recieved_protected_documents_cell._value = str(recieved_protected_documents_name)
                    recieved_policy_cell._value = str(recieved_policy_name)
                    recieved_violation_level_cell.value = str(recieved_violation_level)
                    recieved_policy_type_cell.value = str(recieved_policy_type)
                    recieved_verdict_cell.value = str(recieved_verdict)
                    mark_cell.value = str(mark)
                    first_detection = 1
            recieved_usages_cell = sheet['U' + str(policy_addr)]
            capture_dates_cell = sheet['Y' + str(policy_addr)]
            recieved_usages_cell.value = str(recieved_usages)
            capture_dates_cell.value = str(capture_dates)
        policy_addr = policy_addr + 1
        x = 0
    wb.save(output_table_name)
    print("Done!")
"""

# Запрос на получение каталогов объектов защиты
# url = 'https://192.168.96.130/xapi/event/protectedCatalog?start=0&limit=' + str(NUMBER_OF_EVENTS)
# Запрос на получение событий с определенным каталогом объекта защиты
# url = 'https://192.168.96.130/xapi/event?filter[protected_catalog_id]=4D71F999F23A28BD87E5244A631D84D400000000&start=0&limit=100&'\
# Запрос на получение событий без доп параметров
# url = 'https://192.168.96.130/xapi/event?start=0&limit=' + str(NUMBER_OF_EVENTS)
# Запрос на получение файла, нарушающего политику безопасности
# url = 'https://192.168.103.102/xapi/event/806/raw?downoload=0'
# Запрос на получение текста, нарушающего политику безопасности
# url = 'https://192.168.103.102/xapi/event/815/text'
# Запрос на получение событий с доп параметрами: политикой, каталогом объектов защиты и объектом защиты
# url = 'https://192.168.103.102/xapi/person?start=0&limit=10'
# url = 'https://192.168.103.102/xapi/group'
# try:
#list = sys.argv
#if len(list) < 7:
#    raise WrongInputException()
#input_table_name = list[1]
# output_table_name = "out.xlsx"
#ip_addr = list[2]
token = '1w79w1z40xshocds56p4'
#day = list[4]
#month = list[5]
#year = list[6]
#hour = list[7]
#minute = list[8]
#output_table_name = str(ip_addr + "_" + day + "-" + month + "-" + year + hour + minute + ".xls")
#print("preparing " + ip_addr + "_" + day + "-" + month + "-" + year + hour + minute + ".xls...")
#users_js = LDAP.get_users_js()
#time_tuple = (int(year), int(month), int(day), int(hour), int(minute), 00, 00, 0, 0)
#timestamp = calendar.timegm(time_tuple)
string_timestamp = '1584722120'
HTTP_HEADERS['X-API-Auth-Token'] = token
url = 'https://10.228.6.236:17443/xapi/event?with[protected_documents]&with[policies]&with[protected_catalogs]&with[tags]&with[senders]&with[recipients]&with[recipients_keys]&with[senders_keys]&start=0&limit=' + str(
    NUMBER_OF_EVENTS) + "&filter[date][from]=" + string_timestamp
events = get_events_from_url(url)
print(events)
# create_list_of_events(get_events_from_url(url))
# save_data_to_table(input_table_name, output_table_name)
# table_comparison(input_table_name, output_table_name)
"""
except Exception as e:
    type_name = type(e).__name__
    handled = False
    if type_name == "URLError":
        print("No connection with TM. Be shure that you are wrighting a right IP address")
        handled = True
    if type_name == "WrongInputException":
        print("Wrong input! less than 7 parameters specified.")
        handled = True
    if type_name == "HTTPError":
        code = e.__getattribute__("code")
        if code == 401:
            print("Authorization failed. Maybe wrong token.")
            handled = True
        if code == 403:
            print("Wrong HTTP_HEADERS settings")
            handled = True
        if code == 404:
            print("TM Can't find plugin. Please, check HTTP_HEADERS")
            handled = True
    if type_name == "FileNotFoundError":
        print("Wrong input file name")
        handled = True
    if type_name == "PermissionError":
        print("Exception openning an output file. Close output file")
        handled = True
    if handled == False:
        print("unhandled exception: " + type_name)

    exit()
"""
