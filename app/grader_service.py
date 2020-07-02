from .grader_repository import GraderRepository
from .iwtm_service import IWTMService
from . import utils
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

class GraderService:
    def __init__(self):
        self.grader_repository = GraderRepository()
        self.iwtm_service = IWTMService()
    
    def save_template_file(self, path, file):
        with open(path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
    
    def save_participant_result(self, participant_number, result):
        self.grader_repository.save_participant_result(participant_number, result)
    
    def find_participant_result_by_number(self, number):
        participant_result = self.grader_repository.find_participant_result_by_number(number)
        if not participant_result:
            raise RuntimeError('Participant result not found')
        return participant_result
    
    def save_participant(self, participant):
        self.grader_repository.save_participant(participant)

    def get_all_participants(self):
        return self.grader_repository.get_all_participants()

    def find_participant_by_number(self, number):
        return self.grader_repository.find_participant_by_number(number)

    def save_auth_cookie(self, participant_number, auth_cookie):
        return self.grader_repository.save_auth_cookie(participant_number, auth_cookie)

    def get_auth_cookie(self, participant_number):
        return self.grader_repository.get_auth_cookie(participant_number)

    def load_events(self, iwtm_ip, username, password, date_and_time, template_file_path):
        template_events = self.load_template_events(template_file_path)
        
        auth_cookies = self.iwtm_service.login(iwtm_ip, username, password)
        unique_senders = self.get_unique_senders(template_events)
        unique_recipients = self.get_unique_recipients(template_events)
        iwtm_events = self.iwtm_service.get_parsed_events(iwtm_ip, auth_cookies, date_and_time, unique_senders, unique_recipients)
        
        mapped_events = self.map_events(template_events, iwtm_events)
        protected_objects = self.iwtm_service.get_protected_objects(iwtm_ip, auth_cookies)
        self.iwtm_service.set_object_technologies(mapped_events, protected_objects)
        return mapped_events
    
    def load_template_events(self, filename):
        template_events = []
        wb = load_workbook(filename)
        sheet = wb.active
        row = 3
        while (sheet['D' + str(row)].value != None):
            task_item = {}
            task_item['task_number'] = sheet['A' + str(row)].value
            task_item['policy'] = sheet['B' + str(row)].value
            task_item['protected_object'] = sheet['C' + str(row)].value
            events = []
            current_task = task_item['task_number']
            while (current_task == task_item['task_number'] and sheet['D' + str(row)].value != None):
                event = {}
                event['filename'] = sheet['D' + str(row)].value
                event['sender'] = sheet['E' + str(row)].value
                event['recipient'] = sheet['F' + str(row)].value
                policies = sheet['G' + str(row)].value
                if policies:
                    event['policies'] = policies.split(',')
                else:
                    event['policies'] = []
                protected_objects = sheet['H' + str(row)].value
                if protected_objects:
                    event['protected_objects'] = protected_objects.split(',')
                else:
                    event['protected_objects'] = []
                event['verdict'] = sheet['I' + str(row)].value
                event['violation_level'] = sheet['J' + str(row)].value
                tags = sheet['K' + str(row)].value
                if tags:
                    event['tags'] = tags.split(',')
                else:
                    event['tags'] = []
                events.append(event)
                row += 1
                if sheet['A' + str(row)].value == None:
                    current_task = task_item['task_number']
                else:
                    current_task = sheet['A' + str(row)].value
            task_item['events'] = events
            task_item['stats'] = {}
            template_events.append(task_item)
        return template_events
    
    def get_unique_senders(self, events):
        return self.get_unique_values(events, 'sender')
    
    def get_unique_recipients(self, events):
        return self.get_unique_values(events, 'recipient')

    def get_unique_values(self, events, field):
        result = set()
        for item in events:
            for event in item['events']:
                result.add(event[field])
        return result

    def map_events(self, template_events, iwtm_events):
        mapped_events = []
        for item in template_events:
            mapped_item = item
            for event in mapped_item['events']:
                event['iwtm_event'] = self.find_event_by_id(iwtm_events, event['filename'], event['sender'], event['recipient'])
            mapped_events.append(mapped_item)
        return mapped_events

    def find_event_by_id(self, events, filename, sender, recipient):
        found_event = {}
        for event in events:
            if event['filename'] == filename and event['sender'] == sender and event['recipient'] == recipient:
                found_event = event
                return found_event
        return found_event

    def check_events_normal_mode(self, grouped_events, iwtm_policies):
        self.check_events_max_mode(grouped_events)
        for item in grouped_events:
            self.find_false_triggering_and_update_stats(item, grouped_events, iwtm_policies)
        return grouped_events

    def find_false_triggering_and_update_stats(self, item, grouped_events, iwtm_policies):
        policy = item['policy']
        protected_object = item['protected_object']
        iwtm_policy = self.iwtm_service.find_policy_by_name(iwtm_policies, policy)
        tags = []
        iwtm_violation_level = 'No'
        if iwtm_policy:
            tags = self.iwtm_service.get_tags_from_policy_transfer_rules(iwtm_policy)
            iwtm_violation_level = self.iwtm_service.get_highest_violation_level_from_policy_transfer_rules(iwtm_policy)
        for any_item in grouped_events:
            for policy_event in any_item['events']:
                iwtm_event = policy_event['iwtm_event']
                if (policy in iwtm_event['policies'] and
                        policy not in policy_event['policies']):
                    item['stats']['false_policies'] += 1
                    any_item['stats']['false_policies'] -= 1
                    for tag in tags:
                        if (tag in iwtm_event['tags'] and
                            tag not in policy_event['tags']):
                            item['stats']['false_tags'] += 1
                            any_item['stats']['false_tags'] -= 1
                    if (iwtm_violation_level == iwtm_event['violation_level'] and
                        iwtm_event['violation_level'] != policy_event['violation_level']):
                        item['stats']['wrong_violation_level'] += 1
                        if any_item['stats']['wrong_violation_level'] > 0:
                            any_item['stats']['wrong_violation_level'] -= 1
                    
                    
                if (protected_object in iwtm_event['protected_objects'] and
                        protected_object not in policy_event['protected_objects']):
                    item['stats']['false_objects'] += 1
                    any_item['stats']['false_objects'] -= 1

    def check_events_max_mode(self, grouped_events):
        for item in grouped_events:
            self.init_task_stats(item)
            self.check_events_one_to_one(item)
        return grouped_events

    def check_events_one_to_one(self, item):
        for task_event in item['events']:
            self.find_and_set_event_difference(task_event['iwtm_event'], task_event)
            self.update_task_stats(item, task_event['iwtm_event'], task_event)

    def find_tags_in_task_events(self, item):
        tags = set()
        for event in item['events']:
            for tag in event['tags']:
                tags.add(tag)
        return tags

    def init_task_stats(self, item):
        item['stats']['failed_policies'] = 0
        item['stats']['false_policies'] = 0
        item['stats']['failed_objects'] = 0
        item['stats']['false_objects'] = 0
        item['stats']['wrong_tags'] = 0
        item['stats']['false_tags'] = 0
        item['stats']['wrong_verdict'] = 0
        item['stats']['wrong_violation_level'] = 0

    def get_array_difference(self, first_array, second_array):
        diff = []
        for item in first_array:
            if item not in second_array:
                diff.append(item)
        return diff

    def find_and_set_event_difference(self, iwtm_event, template_event):
        failed_policies_diff = self.get_array_difference(template_event['policies'],
                                                         iwtm_event['policies'])
        false_policies_diff = self.get_array_difference(iwtm_event['policies'],
                                                            template_event['policies'])
        failed_objects_diff = self.get_array_difference(template_event['protected_objects'],
                                                              iwtm_event['protected_objects'])
        false_objects_diff = self.get_array_difference(iwtm_event['protected_objects'],
                                                             template_event['protected_objects'])
        failed_tags_diff = self.get_array_difference(template_event['tags'],
                                                  iwtm_event['tags'])
        false_tags_diff = self.get_array_difference(iwtm_event['tags'],
                                                        template_event['tags'])
        wrong_verdict_diff = ''
        if iwtm_event['verdict'] != template_event['verdict']:
            wrong_verdict_diff = iwtm_event['verdict']
            
        wrong_violation_level_diff = ''
        if iwtm_event['violation_level'] != template_event['violation_level']:
            wrong_violation_level_diff = iwtm_event['violation_level']
            
        iwtm_event['diff'] = {}
        template_event['diff'] = {}
        iwtm_event['diff']['policies'] = false_policies_diff
        template_event['diff']['policies'] = failed_policies_diff
        iwtm_event['diff']['protected_objects'] = false_objects_diff
        template_event['diff']['protected_objects'] = failed_objects_diff
        iwtm_event['diff']['tags'] = false_tags_diff
        template_event['diff']['tags'] = failed_tags_diff
        iwtm_event['diff']['verdict'] = wrong_verdict_diff
        iwtm_event['diff']['violation_level'] = wrong_violation_level_diff
    
    def update_task_stats(self, item, iwtm_event, template_event):
        item['stats']['failed_policies'] += len(template_event['diff']['policies'])
        item['stats']['false_policies'] += len(iwtm_event['diff']['policies'])
        item['stats']['failed_objects'] += len(template_event['diff']['protected_objects'])
        item['stats']['false_objects'] += len(iwtm_event['diff']['protected_objects'])
        item['stats']['wrong_tags'] += len(template_event['diff']['tags'])
        item['stats']['false_tags'] += len(iwtm_event['diff']['tags'])
        if iwtm_event['diff']['verdict']:
            item['stats']['wrong_verdict'] += 1
        if iwtm_event['diff']['violation_level']:
            item['stats']['wrong_violation_level'] += 1

    def export_participant_result(self, result, participant, filename, locale):
        start_x = 1
        start_y = 1
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        book = Workbook()
        sheet = book.active
        x_idx = start_x
        y_idx = start_y

        participant_info_str = 'Участник: ' if locale == 'ru' else 'Competitor: '
        participant_info_str += participant['number'] + ' '
        participant_info_str += participant['last_name'] + ' '
        participant_info_str += participant['ip']
        sheet.cell(x_idx, y_idx).value = participant_info_str
        y_idx += 1
        sheet.cell(x_idx, y_idx).value = 'Метод оценки' if locale == 'ru' else 'Check mode'
        y_idx += 1
        sheet.cell(x_idx, y_idx).value = self.get_check_mode_display_name(participant['check_mode'], locale)
        y_idx = start_y
        x_idx += 1
        
        policy_column_headers = self.get_policy_column_headers(locale)
        
        for header in policy_column_headers:
            sheet.cell(x_idx, y_idx).value = header
            sheet.cell(x_idx, y_idx).border = thin_border
            sheet.cell(x_idx, y_idx).alignment = cell_align
            sheet.column_dimensions[get_column_letter(y_idx)].width = 18
            y_idx += 1
        
        for item in result:
            x_idx += 1
            y_idx = start_y
            policy = item['policy']
            task = item['task_number']
            total_errors = (item['stats']['false_policies'] + item['stats']['failed_policies'] +
                            item['stats']['false_objects'] + item['stats']['failed_objects'] +
                            item['stats']['false_tags'] + item['stats']['wrong_tags'] +
                            item['stats']['wrong_verdict'] + item['stats']['wrong_violation_level'])
            sheet.cell(x_idx, y_idx).value = task
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = policy
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_policies']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_policies']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_tags']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_tags']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_verdict']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_violation_level']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = total_errors
            sheet.cell(x_idx, y_idx).border = thin_border
        
        protected_object_column_headers = self.get_protected_object_column_headers(locale)
        y_idx = start_y
        x_idx += 3
        for header in protected_object_column_headers:
            sheet.cell(x_idx, y_idx).value = header
            sheet.cell(x_idx, y_idx).border = thin_border
            sheet.cell(x_idx, y_idx).alignment = cell_align
            sheet.column_dimensions[get_column_letter(y_idx)].width = 18
            y_idx += 1

        technologies = self.get_technologies_dict(locale)

        for item in result:
            x_idx += 1
            y_idx = start_y
            protected_object = item['protected_object']
            task = item['task_number']
            total_errors = item['stats']['false_objects'] + item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).value = task
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = protected_object
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = total_errors
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            if item['protected_object_technologies']:
                display_names = []
                for technology in item['protected_object_technologies']:
                    display_names.append(self.get_technology_display_name(technologies, technology))
                technologies_str = ", ".join(display_names)
                sheet.cell(x_idx, y_idx).value = technologies_str
            sheet.cell(x_idx, y_idx).border = thin_border
        book.save(filename)

    def get_policy_column_headers(self, locale):
        if locale == 'ru':
            column_headers = [
                'Задание', 'Политика', 'Ложные политики', 'Несработавшие политики', 
                'Ложные объекты', 'Несработавшие объекты', 
                'Ложные теги', 'Неправильные теги', 'Неправильные вердикты',
                'Неправильные уровни нарушения', 'Итого недочетов'
            ]
            return column_headers
        elif locale == 'en':
            column_headers = [
                'Task', 'Policy', 'False policies', 'Failed policies', 
                'False protected objects', 'Failed protected objects', 
                'False tags', 'Wrong tags', 'Wrong verdicts',
                'Wrong violation levels', 'Total errors'
            ]
            return column_headers
        else:
            raise RuntimeError('Unsupported locale')
    
    def get_protected_object_column_headers(self, locale):
        if locale == 'ru':
            column_headers = [
                'Задание', 'Объект защиты', 'Ложные объекты', 
                'Несработавшие объекты', 'Итого недочетов',
                'Типы технологий'
            ]
            return column_headers
        elif locale == 'en':
            column_headers = [
                'Task', 'Protected object', 'False protected objects', 
                'Failed protected objects', 'Total errors',
                'Technologies'
            ]
            return column_headers
        else:
            raise RuntimeError('Unsupported locale')

    def get_technology_display_name(self, technologies, technology):
        display_name = ''
        if technology in technologies:
            display_name = technologies[technology]
        return display_name

    def get_technologies_dict(self, locale):
        if locale == 'ru':
            technologies = {}
            technologies['term'] = 'категория'
            technologies['stamp'] = 'печать'
            technologies['form'] = 'бланк'
            technologies['fingerprint'] = 'эталонный документ'
            technologies['text_object'] = 'текстовый объект'
            technologies['table'] = 'выгрузка из БД'
            technologies['graphic'] = 'графический объект'
            return technologies
        elif locale == 'en':
            technologies = {}
            technologies['term'] = 'category'
            technologies['stamp'] = 'stamp'
            technologies['form'] = 'blank'
            technologies['fingerprint'] = 'sample document'
            technologies['text_object'] = 'text object'
            technologies['table'] = 'DB unloading'
            technologies['graphic'] = 'graphical object'
            return technologies
        else:
            raise RuntimeError('Unsupported locale')

    def get_check_mode_display_name(self, check_mode, locale):
        if locale == 'ru':
            if check_mode == 'normal':
                return 'умеренная оценка'
            else:
                return 'максимальная оценка'
        else:
            if check_mode == 'normal':
                return 'normal check mode'
            else:
                return 'max check mode'