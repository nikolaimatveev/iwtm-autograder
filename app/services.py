import csv
import os
import json
import urllib.parse
import urllib.request
import ssl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import blowfish
import base64
import requests
from datetime import datetime, timezone
from time import gmtime, strftime
from urllib.parse import unquote

class EventService:
    def __init__(self, debug_mode):
        self.debug_mode = debug_mode
        # todo: store in db
        self.participant_results = {}
        self.participants = {}
        self.auth_cookies = []

    def load_events(self, iwtm_ip, iwtm_login, iwtm_password, 
                            date_and_time, template_file_path):
        template_events = self.load_template_events(template_file_path)
        iwtm_events = []
        if self.debug_mode:
            iwtm_events_file = 'app/static/sample-events.json'
            iwtm_events = self.load_iwtm_events_from_file(iwtm_events_file)
        else:
            if not self.isAuthCookiesValid(iwtm_ip, self.auth_cookies):
                self.auth_cookies = self.login_to_iwtm(iwtm_ip, iwtm_login, iwtm_password)
            token = self.get_token_from_iwtm(iwtm_ip, self.auth_cookies)
            iwtm_events = self.load_events_from_iwtm(iwtm_ip, token, date_and_time)
        
        iwtm_events = self.parse_iwtm_events(iwtm_events, iwtm_ip, self.auth_cookies)
        mapped_events = self.map_events(template_events, iwtm_events)
        protected_objects = self.get_protected_objects(iwtm_ip, self.auth_cookies)
        self.set_object_technologies(mapped_events, protected_objects)
        self.save_participant_result(iwtm_ip, mapped_events)
        return True

    def isAuthCookiesValid(self, iwtm_ip, auth_cookies):
        url = 'https://' + iwtm_ip + '/api/user/check'
        response = requests.get(url, cookies=auth_cookies, verify=False)
        data = response.json()
        if response.status_code == 200 and data['state'] == 'data':
            return True
        else:
            return False

    def save_participant(self, ip, participant_info):
        self.participants[ip] = participant_info

    def save_participant_result(self, ip, result):
        self.participant_results[ip] = result

    def get_participants(self):
        return self.participants.values()

    def get_participant_by_ip(self, ip):
        result = {}
        if ip in self.participants:
            result = self.participants[ip]
        return result

    def get_participant_result(self, ip):
        result = {}
        if ip in self.participant_results:
            result = self.participant_results[ip]
        return result
    
    def save_template_file(self, path, file):
        with open(path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
    
    def load_template_events(self, filename):
        template_events = []
        wb = load_workbook(filename)
        sheet = wb.active
        row = 3
        while (sheet['D' + str(row)].value != None):
            task_item = {}
            task_item['task_number'] = sheet['A' + str(row)].value
            task_item['policy'] = sheet['B' + str(row)].value
            task_item['protected_object'] = sheet['C' + str(row)].value
            events = []
            current_task = task_item['task_number']
            while (current_task == task_item['task_number'] and sheet['D' + str(row)].value != None):
                event = {}
                event['filename'] = sheet['D' + str(row)].value
                event['sender'] = sheet['E' + str(row)].value
                event['recipient'] = sheet['F' + str(row)].value
                policies = sheet['G' + str(row)].value
                if policies:
                    event['policies'] = policies.split(',')
                else:
                    event['policies'] = []
                protected_objects = sheet['H' + str(row)].value
                if protected_objects:
                    event['protected_objects'] = protected_objects.split(',')
                else:
                    event['protected_objects'] = []
                event['verdict'] = sheet['I' + str(row)].value
                event['violation_level'] = sheet['J' + str(row)].value
                tags = sheet['K' + str(row)].value
                if tags:
                    event['tags'] = tags.split(',')
                else:
                    event['tags'] = []
                events.append(event)
                row += 1
                if sheet['A' + str(row)].value == None:
                    current_task = task_item['task_number']
                else:
                    current_task = sheet['A' + str(row)].value
            task_item['events'] = events
            task_item['stats'] = {}
            template_events.append(task_item)
        return template_events
    
    def load_iwtm_events_from_file(self, filename):
        data = []
        with open(filename, encoding='utf-8') as json_file:
            data = json.load(json_file)
        return data['data']
    
    def get_token_from_iwtm(self, iwtm_ip, auth_cookies):
        url = 'https://' + iwtm_ip + '/api/plugin?merge_with[]=tokens'
        response = requests.get(url, cookies=auth_cookies, verify=False)
        data = response.json()
        token = ''
        for plugin in data['data']:
            if plugin['DISPLAY_NAME'] == 'DataExport plugin':
                token = plugin['tokens'][0]['USERNAME']
        return token

    def convert_datetime_to_timestamp(self, date_and_time):
        attr = date_and_time.split('-')
        dt = datetime(year=int(attr[0]), month=int(attr[1]), 
                      day=int(attr[2]), hour=int(attr[3]), minute=int(attr[4]))
        timestamp = dt.replace(tzinfo=timezone.utc).timestamp()
        return timestamp

    def load_events_from_iwtm(self, iwtm_ip, token, date_and_time):
        url = 'https://' + iwtm_ip + '/xapi/event'
        headers = {}
        headers['X-API-Version'] = '1.2'
        headers['X-API-CompanyId'] = 'GUAP'
        headers['X-API-ImporterName'] = 'GUAP'
        headers['X-API-Auth-Token'] = token
        params = {}
        params['start'] = 0
        timestamp = self.convert_datetime_to_timestamp(date_and_time)
        params['filter[date][from]'] = timestamp
        params['with[]'] = [
            'protected_documents', 'policies', 'protected_catalogs',
            'tags', 'senders', 'recipients',
            'senders_keys', 'recipients_keys',
            'perimeters', 'attachments']
        response = requests.get(url, headers=headers, params=params, verify=False)
        data = response.json()
        if data['meta']['totalCount'] == 0:
            raise RuntimeError('No events found for the specified time period')
        return data['data']

    def login_to_iwtm(self, iwtm_ip, username, password):
        api_root = 'https://' + iwtm_ip + '/api/'
        resp = requests.get(api_root + 'user/check', verify=False)
        csrf_token = unquote(resp.cookies['YII_CSRF_TOKEN'])
    
        response = requests.get(api_root + 'salt', 
                                cookies=resp.cookies, verify=False)
    
        response_json = response.json()
        salt = response_json['data']['salt']
        crypted_password = self.crypt_password(password, salt)
    
        data = {}
        data['username'] = username
        data['crypted_password'] = crypted_password

        headers = {}
        headers['x-csrf-token'] = csrf_token
    
        response = requests.post(api_root + 'login',
                                json=data, cookies=resp.cookies, 
                                headers=headers, verify=False)
        return response.cookies
    
    def crypt_password(self, password, salt):
        key_bytes = salt.encode('utf-8')
        cipher = blowfish.Cipher(key_bytes)
        msg_bytes = password.encode('utf-8')
        
        block_size = 8
        padding_len = -len(msg_bytes) % block_size
        msg_bytes = msg_bytes.ljust(len(msg_bytes) + padding_len, b'\0')
        
        data_encrypted = b"".join(cipher.encrypt_ecb(msg_bytes))
        crypted_bytes = base64.b64encode(data_encrypted)
        crypted_base64 = crypted_bytes.decode('utf-8')
        return crypted_base64
    
    def get_protected_objects(self, iwtm_ip, auth_cookies):
        url = 'https://' + iwtm_ip + '/api/protectedDocument?start=0&limit=100'
        response = requests.get(url, cookies=auth_cookies, verify=False)
        data = response.json()
        return data['data']

    def find_protected_object_by_name(self, protected_objects, name):
        found_objects = []
        for protected_object in protected_objects:
            if name in protected_object['DISPLAY_NAME']:
                found_objects.append(protected_object)
        return found_objects
    
    def get_protected_object_technologies(self, protected_objects):
        technologies = set()
        for protected_object in protected_objects:
            for entry in protected_object['entries_pool']:
                if entry['ENTRY_TYPE'] == 'text_object':
                    technologies.add(entry['ENTRY_TYPE'])
                else:
                    technology = entry['content']['TYPE']
                    technologies.add(technology)
        return technologies
    
    def set_object_technologies(self, events, protected_objects):
        for item in events:
            protected_object_name = item['protected_object']
            technologies = []
            if protected_object_name:
                found_protected_objects = self.find_protected_object_by_name(protected_objects, protected_object_name)
                technologies = self.get_protected_object_technologies(found_protected_objects)
            item['protected_object_technologies'] = technologies
    
    def parse_iwtm_events(self, events, iwtm_ip, auth_cookies):
        parsed_events = []
        for event in events:
            parsed_event = {}
            parsed_event['id'] = event['OBJECT_ID']
            parsed_event['capture_date'] = event['CAPTURE_DATE']
            subject = json.loads(event['PREVIEW_DATA'])['subject'].split(', ')
            event_sender = subject[0].split(':')[1].strip()
            event_recipient = subject[1].split(':')[1].strip()
            event_filename = subject[2].split(':')[1].strip()
            parsed_event['sender'] = event_sender
            parsed_event['recipient'] = event_recipient
            parsed_event['filename'] = event_filename
            parsed_event['policies'] = self.simplify_json_array(event['policies'])
            parsed_event['protected_objects'] = self.simplify_json_array(event['protected_documents'])
            parsed_event['tags'] = self.simplify_json_array(event['tags'])
            parsed_event['verdict'] = event['VERDICT']
            parsed_event['violation_level'] = event['VIOLATION_LEVEL']
            parsed_events.append(parsed_event)
        return parsed_events
    
    def map_events(self, template_events, iwtm_events):
        mapped_events = []
        for item in template_events:
            mapped_item = item
            for event in mapped_item['events']:
                event['iwtm_event'] = self.find_event_by_id(iwtm_events, event['filename'], event['sender'], event['recipient'])
            mapped_events.append(mapped_item)
        return mapped_events

    def find_event_by_id(self, events, filename, sender, recipient):
        found_event = {}
        for event in events:
            if event['filename'] == filename and event['sender'] == sender and event['recipient'] == recipient:
                found_event = event
                return found_event
        return found_event

    def simplify_json_array(self, json_array):
        result = []
        for item in json_array:
            result.append(item['DISPLAY_NAME'])
        return result

    def check_events_normal_mode(self, grouped_events):
        self.check_events_max_mode(grouped_events)
        for item in grouped_events:
            self.find_false_triggering_and_update_stats(item, grouped_events)
        return grouped_events

    def find_false_triggering_and_update_stats(self, item, grouped_events):
        policy = item['policy']
        protected_object = item['protected_object']
        for any_item in grouped_events:
            for policy_event in any_item['events']:
                iwtm_event = policy_event['iwtm_event']
                if (policy in iwtm_event['policies'] and
                        policy not in policy_event['policies']):
                    item['stats']['false_policies'] += 1
                    any_item['stats']['false_policies'] -= 1
                    item['stats']['false_tags'] += 1
                    any_item['stats']['false_tags'] -= 1
                    
                if (protected_object in iwtm_event['protected_objects'] and
                        protected_object not in policy_event['protected_objects']):
                    item['stats']['false_objects'] += 1
                    any_item['stats']['false_objects'] -= 1

    def check_events_max_mode(self, grouped_events):
        for item in grouped_events:
            self.init_task_stats(item)
            self.check_events_one_to_one(item)
        return grouped_events

    def check_events_one_to_one(self, item):
        for task_event in item['events']:
            self.find_and_set_event_difference(task_event['iwtm_event'], task_event)
            self.update_task_stats(item, task_event['iwtm_event'], task_event)

    def init_task_stats(self, item):
        item['stats']['failed_policies'] = 0
        item['stats']['false_policies'] = 0
        item['stats']['failed_objects'] = 0
        item['stats']['false_objects'] = 0
        item['stats']['wrong_tags'] = 0
        item['stats']['false_tags'] = 0
        item['stats']['wrong_verdict'] = 0
        item['stats']['wrong_violation_level'] = 0

    def get_array_difference(self, first_array, second_array):
        diff = []
        for item in first_array:
            if item not in second_array:
                diff.append(item)
        return diff

    def find_and_set_event_difference(self, iwtm_event, template_event):
        failed_policies_diff = self.get_array_difference(template_event['policies'],
                                                         iwtm_event['policies'])
        false_policies_diff = self.get_array_difference(iwtm_event['policies'],
                                                            template_event['policies'])
        failed_objects_diff = self.get_array_difference(template_event['protected_objects'],
                                                              iwtm_event['protected_objects'])
        false_objects_diff = self.get_array_difference(iwtm_event['protected_objects'],
                                                             template_event['protected_objects'])
        failed_tags_diff = self.get_array_difference(template_event['tags'],
                                                  iwtm_event['tags'])
        false_tags_diff = self.get_array_difference(iwtm_event['tags'],
                                                        template_event['tags'])
        wrong_verdict_diff = ''
        if iwtm_event['verdict'] != template_event['verdict']:
            wrong_verdict_diff = iwtm_event['verdict']
            
        wrong_violation_level_diff = ''
        if iwtm_event['violation_level'] != template_event['violation_level']:
            wrong_violation_level_diff = iwtm_event['violation_level']
            
        iwtm_event['diff'] = {}
        template_event['diff'] = {}
        iwtm_event['diff']['policies'] = false_policies_diff
        template_event['diff']['policies'] = failed_policies_diff
        iwtm_event['diff']['protected_objects'] = false_objects_diff
        template_event['diff']['protected_objects'] = failed_objects_diff
        iwtm_event['diff']['tags'] = false_tags_diff
        template_event['diff']['tags'] = failed_tags_diff
        iwtm_event['diff']['verdict'] = wrong_verdict_diff
        iwtm_event['diff']['violation_level'] = wrong_violation_level_diff
    
    def update_task_stats(self, item, iwtm_event, template_event):
        item['stats']['failed_policies'] += len(template_event['diff']['policies'])
        item['stats']['false_policies'] += len(iwtm_event['diff']['policies'])
        item['stats']['failed_objects'] += len(template_event['diff']['protected_objects'])
        item['stats']['false_objects'] += len(iwtm_event['diff']['protected_objects'])
        item['stats']['wrong_tags'] += len(template_event['diff']['tags'])
        item['stats']['false_tags'] += len(iwtm_event['diff']['tags'])
        if iwtm_event['diff']['verdict']:
            item['stats']['wrong_verdict'] += 1
        if iwtm_event['diff']['violation_level']:
            item['stats']['wrong_violation_level'] += 1
    
    def export_participant_result(self, result, filename):
        start_x = 2
        start_y = 1
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        book = Workbook()
        sheet = book.active
        x_idx = start_x
        y_idx = start_y
        column_headers = ['Задание', 'Политика', 'Ложные политики', 'Несработавшие политики', 
                        'Ложные объекты', 'Несработавшие объекты', 
                        'Ложные теги', 'Неправильные теги', 'Неправильные вердикты',
                        'Неправильные уровни нарушения', 'Итого недочетов']
        
        for header in column_headers:
            sheet.cell(x_idx, y_idx).value = header
            sheet.cell(x_idx, y_idx).border = thin_border
            sheet.cell(x_idx, y_idx).alignment = cell_align
            sheet.column_dimensions[get_column_letter(y_idx)].width = 18
            y_idx += 1
        
        for item in result:
            x_idx += 1
            y_idx = start_y
            policy = item['policy']
            task = item['task_number']
            total_errors = (item['stats']['false_policies'] + item['stats']['failed_policies'] +
                            item['stats']['false_objects'] + item['stats']['failed_objects'] +
                            item['stats']['false_tags'] + item['stats']['wrong_tags'] +
                            item['stats']['wrong_verdict'] + item['stats']['wrong_violation_level'])
            sheet.cell(x_idx, y_idx).value = task
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = policy
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_policies']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_policies']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_tags']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_tags']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_verdict']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['wrong_violation_level']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = total_errors
            sheet.cell(x_idx, y_idx).border = thin_border
        
        
        column_headers_doc = ['Задание', 'Объект защиты', 'Ложные объекты', 
                              'Несработавшие объекты', 'Итого недочетов', 'Типы технологий']
        y_idx = start_y
        x_idx += 3
        for header in column_headers_doc:
            sheet.cell(x_idx, y_idx).value = header
            sheet.cell(x_idx, y_idx).border = thin_border
            sheet.cell(x_idx, y_idx).alignment = cell_align
            sheet.column_dimensions[get_column_letter(y_idx)].width = 18
            y_idx += 1

        for item in result:
            x_idx += 1
            y_idx = start_y
            protected_object = item['protected_object']
            task = item['task_number']
            total_errors = item['stats']['false_objects'] + item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).value = task
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = protected_object
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['false_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = item['stats']['failed_objects']
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            sheet.cell(x_idx, y_idx).value = total_errors
            sheet.cell(x_idx, y_idx).border = thin_border
            y_idx += 1
            if item['protected_object_technologies']:
                technologies_str = ", ".join(item['protected_object_technologies'])
                sheet.cell(x_idx, y_idx).value = technologies_str
            sheet.cell(x_idx, y_idx).border = thin_border
        book.save(filename)